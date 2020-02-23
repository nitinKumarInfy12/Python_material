#! python3
# openpyxl module allows python program to read/wrtite excel spreadsheeets
# workbook : is the excel document , saved as .xlsx
# worksheet : the individual tabs in the excel
# active sheet : the worksheet user is currently using or last viedwed before closing the excel
# each sheet has cloumn starting from A, and Rows starting from 1
# Cell : a box at a particular intersection. The grid of cells with data makes up a sheet
# openpyxl.load_workbook()   : this takes the file name and returns object of workbook
# sheetnames attribute  returns all the sheet names in the excel document. use worksheetobject.sheetnames
# worksheetobject[SheetName]  retruns the worksheet object of the provided sheet. use worksheetobject[SheetName]
# title attribute returns the sheet name
# once the worksheet object is cerated, cell objects can be accessed bu its name
# value, row, column, coordinate attributes of the cell object can be used to get the cell info
# sheet.cell() method on sheet is better choice than Sheet[Cell Name]
# sheet.cell(row=1, column=2) gets you a Cell object for cell B1, just like specifying sheet['B1']
# by using cell(). loops can be used easily over the sheet object
# Size of the sheet an be calculated by using get_highest_row() and get_highest_column() methods
# max_row/ max_column properties retuns maximum row/column index number like 1,2,3 etc.
# openpyxl.utils.column_index_from_string() and openpyxl.utils.get_column_letter() function can be used to convert cell index to name string or vice versa
# To convert from letters to numbers use :openpyxl.utils.column_index_from_string() function. 
# To convert from numbers to letters use : openpyxl.cell.get_column_letter() function.
# using SheetObject[A1:C3] worksheet can be sliced in rows, colums or rectangle
# SheetObject[A1:c3]  returns each cell objects in rows and each row has the cell objects of the row
# similarly Worksheet objects rows and columns attributes can be used to get tuples of cell objects from the rows and columns
# indexes can be used along with rows/column attributes to get a particular colum, or index
# sheetobject.rows   will retun the tuple of cell objects in rows, each row will have the cell objects for each colum.
# sheetobject[2]  will return the tuple of cell objects for each column/cell in the 2nd row of the worksheet.
# sheetobject.columns   will retun the tuple of cell objects in columns, each column tuple will have the cell objects for each row.
# sheetobject['B']  will return the tuple of cell objects for each row/cell in the 2nd column of the worksheet.


# summary: read the execl
As a quick review, here’s a rundown of all the functions, methods, and data types involved in reading a cell out of a spreadsheet file:
1. Import the openpyxl module.
2. Call the openpyxl.load_workbook() function to load the existing workbook.
3. Get a Workbook object.
4. Call the wb[sheetname] workbook attribute.
5. Get a Worksheet object.
6. Use indexing or the cell() sheet method with row and column keyword arguments.
7. Get a Cell object.
8. Read the Cell object’s value attribute
 
 
# summary: create/write the execl
As a quick review, here’s a rundown of all the functions, methods, and data types involved in reading a cell out of a spreadsheet file:
1. Import the openpyxl module.
2. call openpyxl.Workbook() function to create a new workbook.
3. it starts with a single sheet "Sheet". it can be renamed with sheet.title  = 'new_name'
4. Get a Workbook object.
4. Call the wb[sheetname] workbook attribute.
5. Get a Worksheet object.
6. wb.save()   :to save any changes to workbook. wb.save("example.xlsx")
7. Passing a different filename than the original, such as 'example_copy.xlsx', saves the changes to a copy of the spreadsheet
8. create_sheet() / create_sheet(index= , title=) to create sheets in the workbook
9. remove_sheet()  to delete the sheets from the workbook. this function takes the sheet object not the sheet name
9. wb.remove_sheet(wb['sheet1'])   or wb.remove(wb[sheetname]) or del wb[shetname]
10. writing to cell is similar to writing in dictionary.ex  sheetObj['A1'] = 'Hello world!'
11. Writing formula to cell sheet['c1'] = '=sum(A1:B1)'
12. data_only=True :to see the result of teh calculation than the actual formula in teh cell data_only=True must be passed as second arhumenyt in load_workbook()


# font and style can be changed by using Font() and Style() functions from "from openpyxl.styles import Font, Style"
# name, size, bold, italic are teh parameters can be passed to Font() as an options
sheet = wb.get_sheet_by_name('Sheet')
italic24Font = Font(size=24, italic=True)
styleObj = Style(font=italic24Font)
sheet['A'].style/styleObj

# adjust row column height and width using row_dimensions and column_dimensions attributes
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
 
#row_dimensions : take numbers as parameter 
#column_dimensions : take Alphabets as paramater
 
 # hide coumn :
 sheet.column_dimensions['C'].hidden = True
 
# cells can be merged / unmerged in sheet by using merge_cells() or unmerger_cells()
sheet.merge_cells('C5:D5')
sheet.unmerge_cells('C5:D5')

# freeze panes can be achived using freeze_panes attribute of teh worksheet object
# To unfreeze all panes, set freeze_panes to None or 'A1'
# it freezes all left columns and all above rows from  cell where free_pane is set.
sheet.freeze_panes = 'A2'  # freezes Row 1
sheet.freeze_panes = 'B1' # freezes Column A
sheet.freeze_panes = 'C1' # freezes Columns A and B
sheet.freeze_panes = 'C2' # freezes Row 1 and columns A and B
sheet.freeze_panes = 'A1'   # No frozen panes
sheet.freeze_panes = None  # No frozen panes


# chart in excel
"""
OpenPyXL supports creating bar, line, scatter, and pie charts using the data in a sheet’s
cells. To make a chart, you need to do the following:
1. create workbook object
2. create worksheet object
1. Create a Reference object from a rectangular selection of cells by using openpyxl.chart.Reference() function by passing 3 parameters to it:
    ex: refObj = openpyxl.chart.Reference(sheet, min_row = 1, min_col = 1, max_row = 10, max_col = 1)
    a. workseet object that has the chart data
    b. tuple of 2 integers, telling rows, colums of top-left cell of the chart data grid
    c. tuple of 2 integers, telling rows, colums of bottom-right cell of the chart data grid
2. Create a Series object by passing in the Reference object.
3. Create a Chart object.
4. Append the Series object to the Chart object.
5. Optionally, set the drawing.top, drawing.left, drawing.width, and
drawing.height variables of the Chart object.
6. Add the Chart object to the Worksheet object.
"""


 
 
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Style

================= reading excel==================
wb = openpyxl.load_workbook('example.xlsx')
type(wb)  # <class 'openpyxl.workbook.workbook.Workbook'>

wb.sheetnames # ['Sheet1','Sheet2','Sheet3']
sheet = wb.sheetname['Sheet3']  # <worksheet Sheet3>
type(sheet)  # <class 'openpyxl.worksheet.worksheet.Worksheet'>

print(sheet.title) # Sheet3

activeSheet = wb[Sheet1]  #< worksheet Sheet1>

c = activeSheet['B1']
type(c) # <Cell Sheet1.B1>

c.value  # Apples
c.row # 1
c.column # B
c.coordinate # B1

activeSheet.cell(row=1, column=2) #<Cell Sheet1.B1>
activeSheet.cell(row=1, column=2).value  #'Apples'


for i in range(1, 8, 2):
    print(i, activeSheet.cell(row=i, column=2).value)

activeSheet.max_column   # 3
activeSheet.max_row      # 7

get_column_letter(activeSheet.max_column)  # C

column_index_from_string('B')  # 2
column_index_from_string('AA')  # 27

tuple(activeSheet['A1':'C3'])
"""((<Cell Sheet1.A1>, <Cell Sheet1.B1>, <Cell Sheet1.C1>), (<Cell Sheet1.A2>,
<Cell Sheet1.B2>, <Cell Sheet1.C2>), (<Cell Sheet1.A3>, <Cell Sheet1.B3>,
<Cell Sheet1.C3>)) """

for rowofCellObjects in activeSheet['A1':'C3']:
    for cellObject in rowofCellObjects:
        print(cellObject.coordinate, cellObject.value)
    print("End of rows)
    
"""
A1 2015-04-05 13:34:02
B1 Apples
C1 73
--- END OF ROW ---
A2 2015-04-05 03:41:23
B2 Cherries
C2 85
--- END OF ROW ---
A3 2015-04-06 12:46:51
B3 Pears
C3 14
--- END OF ROW ---
"""

activeSheet.columns[1] 
# (<Cell Sheet1.B1>, <Cell Sheet1.B2>, <Cell Sheet1.B3>, <Cell Sheet1.B4>,<Cell Sheet1.B5>, <Cell Sheet1.B6>, <Cell Sheet1.B7>)
# for cellObj in activeSheet.columns[1]:   # this does not work in new version
#for cellObj in activeSheet['B']:  # this is teh working command or teh following one in new version
for cellObj in list(activeSheet.columns)[1]:
    print(cellObj.value)
"""    
Apples
Cherries
Pears
Oranges
Apples
Bananas
Strawberries
"""

====================== writing excel=============================

wb = openpyxl.Workbook()
sheet = wb['Sheet']
#print(sheet)

print(sheet.title)
# update teh sheet name
sheet.title = 'Test Sheet'

print(sheet.title)
print(wb.sheetnames)

# save teh changes done to woekbook use wb.save(workbok name) # giving a different name will create a new copy with chnages savid in.
#wb.save("file_name.xslx")

# create sheets in the workbook use create_sheet() / create_sheet(index= , title= )
wb.create_sheet('mysheet1')
print(wb.sheetnames)
print(wb['mysheet1'])

wb.create_sheet(index=1, title='mysheet2')
print(wb.sheetnames)

wb.create_sheet(index=0, title='mysheet2')
print(wb.sheetnames)

# delete sheets from teh workbook remove_sheet(sheetobject to be deleted)   this function takes the sheet object as parameter not the string name
wb.remove(wb['mysheet21'])
print(wb.sheetnames)

# writing a value in cell
print(sheet.title)
sheet['A2'] = 'Hello Test val'

# writing formula to cell

sheet['c1'] = '=sum(A1:B1)'

# save changes to the file
wb.save('test.xlsx')



# ============= charts=====================================================
# created a bar chart by calling openpyxl.charts.BarChart().
# You can also create line charts, scatter charts, and pie charts by calling openpyxl.charts.LineChart(),
# openpyxl.charts.ScatterChart(), and openpyxl.charts.PieChart()
# load_workbook or Workbook methods dont load the chart in the object, even if there is a chart in teh source excel.
# it can only create chart but cannot read chart.

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11): # create some data in column A
    sheet['A' + str(i)] = i
    #print(f"first i: {i}")

for i in range(10, 0,-1): # create some data in column A
    sheet['B' + str(10-(i-1))] = i
    #print(f"Second i: {i}")

""" single series chart
#refObj = openpyxl.chart.Reference(sheet, (1, 1), (10, 1))
refObj1 = openpyxl.chart.Reference(sheet, min_row = 1, min_col = 1, max_row = 10, max_col = 1)
seriesObj1 = openpyxl.chart.Series(refObj1, title='First series')
chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj1)

#chartObj.drawing.top = 50 # set the position
#chartObj.drawing.left = 100
#chartObj.drawing.width = 300 # set the size
#chartObj.drawing.height = 200
sheet.add_chart(chartObj)
wb.save('sampleBarChart.xlsx')

"""
# multiple series chart
refObj1 = openpyxl.chart.Reference(sheet, min_row = 1, min_col = 1, max_row = 10, max_col = 1)
seriesObj1 = openpyxl.chart.Series(refObj1, title='First series')

refObj2 = openpyxl.chart.Reference(sheet, min_row = 1, min_col = 2, max_row = 10, max_col = 2)
seriesObj2 = openpyxl.chart.Series(refObj2, title='Second series')

chartObj = openpyxl.chart.BarChart()   # creates teh bar cart object
chartObj.title = 'My Chart'

# append series objects to chart
chartObj.append(seriesObj1)
chartObj.append(seriesObj2)

sheet.add_chart(chartObj)
wb.save('sampleBarChart.xlsx')







