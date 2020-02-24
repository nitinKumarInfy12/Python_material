#! python3

# this prject fetches the pyhton projects from the guthub public API
# Git hub Public API : https://api.github.com/search/repositories?q=language:python&sort=stars
# https://api.github.com/rate_limit   to find out the search limit per minute

import requests
import json
import openpyxl
import os

url = 'https://api.github.com/search/repositories?q=language:python&sort=stars'
headers = {'Accept': 'application/vnd.github.v3+json'}

# delete the excel file if esists
filename = '.\\Python_GitHub_Projects.xlsx'
if os.path.exists(filename):
    os.unlink(filename)

# create the response object
req = requests.get(url, headers=headers)
print(req.status_code)
req.raise_for_status()

# convert into python dictionary
json_str = req.text
git_dict = json.loads(json_str)
print(type(git_dict))

# convert into python dictionary : shortcut
git_dict_sc = req.json()
print(type(git_dict_sc))

# print the keys
print(git_dict_sc.keys())
print(f"Total reositories : {git_dict_sc['total_count']}")

repo_dicts = git_dict_sc['items']
print(f"Repositories returned: {len(repo_dicts)}")

# Examine the first repository.
#repo_dict = repo_dicts[0]
#print(f"\nKeys: {len(repo_dict)}")
#for key in sorted(repo_dict.keys()):
#    print(key)

print("\nSelected information about first repository:")

# create excel objects
wb = openpyxl.Workbook()
ws = wb['Sheet']
i=1
#for repo_dict in repo_dicts:
for k in range(len(repo_dicts)):
    repo_dict = repo_dicts[k]
    i = k +1
    if i>1:
        ws.cell(row=i, column=1).value = repo_dict['name']
        ws.cell(row=i, column=2).value = repo_dict['owner']['login']
        ws.cell(row=i, column=3).value = repo_dict['stargazers_count']
        ws.cell(row=i, column=4).value = repo_dict['created_at']
        ws.cell(row=i, column=5).value = repo_dict['updated_at']
        ws.cell(row=i, column=6).value = repo_dict['html_url']
        ws.cell(row=i, column=7).value = repo_dict['description']

    else:
        ws.cell(row=1, column=1).value = 'Project Name'
        ws.cell(row=1, column=2).value = 'Owner Name'
        ws.cell(row=1, column=3).value = 'Start_Count'
        ws.cell(row=1, column=4).value = 'Created on'
        ws.cell(row=1, column=5).value = 'Updated On'
        ws.cell(row=1, column=6).value = 'URL'
        ws.cell(row=1, column=7).value = 'Description'



wb.save(filename)




