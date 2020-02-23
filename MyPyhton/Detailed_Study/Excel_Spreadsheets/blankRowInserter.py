#! python3

"""Replicate a spreadsheet in CWD but with N blank rows inserted at chosen place.

Usage: blank_row_inserter.py takes 3 arguments
<Name> - Name of the spreadsheet to add blank rows to
<Start> - The row to start inserting the blank lines
<Number> - The number of blank lines to insert

e.g. python blank_row_inserter.py test.xlsx 3 2
     Would Insert 2 blank rows starting at row 3 into a copy of 'test.xlxs'
"""

import openpyxl

startNum = 3
endNum = 2
fileName = 'myProduce.xlsx'

wb = openpyxl.load_workbook(fileName)
#sheet = wb.sheetnames
#print(sheet)
ws = wb.active
#ws =wb['Sheet']

SheetName = ws.title

print(ws)
print(SheetName)

wb.create_sheet('Sheet_Copy')
#print(wb.sheetnames)

ws_Copy = wb['Sheet_Copy']
#print(ws_Copy)
#print(ws.max_row, ws.max_column)

for line in range(1, ws.max_row+1):
    for column in range(1,ws.max_column+1):
        if(line >=startNum):
            ws_Copy.cell(row=(line+endNum), column=column).value = ws.cell(row=line, column=column).value
        else:
            ws_Copy.cell(row=line, column=column).value = ws.cell(row=line, column=column).value


# delete the old sheet
wb.remove(ws)

# rename the new sheet to old sheet name
ws_Copy.title = SheetName

wb.save(fileName)
print("Done Inserting blank rows")




