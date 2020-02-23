#! python3

# this program copies data from text files into excel columns

import re, os, openpyxl
from openpyxl.styles import Font

res = re.compile(r'.*(txt)')

path = 'C:\\Users\\kumar\\PycharmProjects\\Excel_Spreadsheets'
excelFile = 'Excel_Notepad.xlsx'


wb = openpyxl.Workbook()
ws= wb.active
bold = Font(bold=True)

Allfiles = os.listdir()
print(Allfiles)
txtFiles = []

for file in Allfiles:
    if res.search(file):
        txtFiles.append(file)

print(txtFiles)
columnIndex = 1

for file in txtFiles:
    sourceTxtFile = f"{path}\\{file}"
    columnHeaderName = file.split('.')[0].upper()
    #print(columnHeaderName)

    with open(sourceTxtFile) as file:
        fileData = file.readlines()

        ws.cell(row=1, column=columnIndex).value = columnHeaderName
        ws.cell(row=1, column=columnIndex).font = bold
        for line in range(1, len(fileData)+1):
            ws.cell(row=line+1, column=columnIndex).value = fileData[line-1]

    columnIndex += 1

wb.save(excelFile)
print("Done adding data")