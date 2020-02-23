#! python3
import openpyxl

# Step 1: Set Up a Data Structure with the Update Information

    # The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}
#print(PRICE_UPDATES.keys())
# create  instances of the workbook and worksheet
wb = openpyxl.load_workbook('produceSales.xlsx')
# print(wb.sheetnames)
sheet = wb['Sheet']

# TODO: Loop through the rows and update the prices.
for rowobjects in sheet['A1':'B'+str(sheet.max_row)]:
    #print(rowobjects[0].value, rowobjects[1].value)
    if rowobjects[0].value in PRICE_UPDATES.keys():
        print(f"{rowobjects[0].value} value in file is {rowobjects[1].value} and in dictionary is {PRICE_UPDATES[rowobjects[0].value]}")
        rowobjects[1].value= PRICE_UPDATES[rowobjects[0].value]
        print("Newvalue")
        print(f"{rowobjects[0].value} new value in file is {rowobjects[1].value} and in dictionary is {PRICE_UPDATES[rowobjects[0].value]}")

wb.save('produceSales.xlsx')
print("Done.")





