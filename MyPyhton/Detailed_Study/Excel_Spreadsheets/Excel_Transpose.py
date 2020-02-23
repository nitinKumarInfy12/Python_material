#! python3

# This program transpose teh rows and columns

import openpyxl

fileName = "Transpose.xlsx"
wb = openpyxl.load_workbook(fileName)
ws = wb.active

sheeetName  = ws.title
#print(sheeetName)

newSheet = 'Sheet_Copy'
wb.create_sheet(newSheet)

wsCopy = wb[newSheet]
#print(wsCopy)
#print(wb.sheetnames)
#print(wsCopy.title)

for x in range(1, ws.max_row+1):
    for y in range(1, ws.max_column+1):
        wsCopy.cell(row=y, column=x).value =ws.cell(row=x, column=y).value

wb.remove(ws)
wsCopy.title = sheeetName

wb.save(fileName)
print("Done transposing the data")



