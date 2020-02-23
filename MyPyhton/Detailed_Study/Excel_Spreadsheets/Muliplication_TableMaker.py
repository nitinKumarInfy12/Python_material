import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
#x = int(input("Input teh integer number: "))
x = 5
wb = openpyxl.Workbook()
sheet = wb.active
#sheet = wb.sheetnames
#sheet = wb['Sheet']
#print(sheet)

sheet.title = 'Muti_table'
#print(wb.sheetnames)
#print(sheet)
make_bold = Font(bold=True)

for i in range(x+1):
#    if sheet['A'+str(i+1)] == sheet['A0']:
#        sheet[get_column_letter(i + 2) + str(1)] = i + 1
#    else:
    sheet['A'+str(i+1)] = i
    sheet['A'+str(i+1)].font = make_bold
    sheet[get_column_letter(i + 2) + str(1)] = i + 1
    sheet[get_column_letter(i + 2) + str(1)].font = make_bold

for i in range(x):
    for y in range(x + 1):
        print(f"{sheet.cell(row=int(i+2), column=int(y+2)) }=A{str(i+2)} * {get_column_letter(int(y+2))}1")
        sheet.cell(row=int(i+2), column=int(y+2)).value = f"=A{str(i + 2)} * {get_column_letter(int(y + 2))}1"

wb.save('Multipplication_table1.xlsx')


